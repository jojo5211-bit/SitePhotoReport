from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from PIL import Image

from app.exporters import ExportSettings, export_excel, export_pdf, export_word, project_info_lines, unique_output_path
from app.settings import AppSettings
from app.store import ProjectStore


class ProjectStoreTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name) / "測試工程.sprj"
        self.store = ProjectStore(self.root)
        self.store.ensure_default_sections()
        self.source_dir = Path(self.temp.name) / "source"
        self.source_dir.mkdir()
        for index, color in enumerate(((220, 50, 50), (50, 160, 80), (50, 80, 220))):
            Image.new("RGB", (640, 480), color).save(self.source_dir / f"photo_{index}.jpg")

    def tearDown(self):
        self.store.close()
        self.temp.cleanup()

    def test_import_and_manual_relocation(self):
        imported, skipped = self.store.import_files(sorted(self.source_dir.glob("*.jpg")))
        self.assertEqual(3, len(imported))
        self.assertEqual([], skipped)
        self.assertEqual(3, len(self.store.placements_for_slot(None)))
        section = self.store.sections()[1]
        slot = self.store.slots(section["id"])[0]
        self.store.relocate_photos(imported[:2], None, slot["id"])
        self.assertEqual(1, len(self.store.placements_for_slot(None)))
        self.assertEqual(2, len(self.store.placements_for_slot(slot["id"])))
        self.store.reorder_placements(slot["id"], list(reversed(imported[:2])))
        ordered = [row["photo_id"] for row in self.store.placements_for_slot(slot["id"])]
        self.assertEqual(list(reversed(imported[:2])), ordered)

    def test_multiple_photos_can_move_between_fields(self):
        imported, _ = self.store.import_files(sorted(self.source_dir.glob("*.jpg")))
        first_section = self.store.sections()[0]
        second_section = self.store.sections()[1]
        source_slot = self.store.slots(first_section["id"])[0]
        target_slot = self.store.slots(second_section["id"])[0]
        self.store.relocate_photos(imported, None, source_slot["id"])
        self.store.relocate_photos(imported[:2], source_slot["id"], target_slot["id"])
        self.assertEqual(1, len(self.store.placements_for_slot(source_slot["id"])))
        self.assertEqual(2, len(self.store.placements_for_slot(target_slot["id"])))

    def test_duplicate_hash_is_skipped(self):
        source = self.source_dir / "photo_0.jpg"
        first, _ = self.store.import_files([source])
        second, skipped = self.store.import_files([source])
        self.assertEqual(1, len(first))
        self.assertEqual([], second)
        self.assertEqual([str(source)], skipped)

    def test_export_formats(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg"])
        section = self.store.sections()[0]
        slot = self.store.slots(section["id"])[0]
        self.store.relocate_photos(imported, None, slot["id"])
        placement = self.store.placement_for_photo(imported[0], slot["id"])
        self.store.update_placement(placement["id"], caption="測試照片說明", construction_location="A棟屋頂")
        pdf = export_pdf(self.store, self.root / "exports" / "report.pdf")
        word = export_word(self.store, self.root / "exports" / "report.docx")
        excel = export_excel(self.store, self.root / "exports" / "report.xlsx")
        for path in (pdf, word, excel):
            self.assertTrue(path.exists())
            self.assertGreater(path.stat().st_size, 0)

    def test_pdf_has_full_cover_before_photo_chapters(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg"])
        self.store.update_project(project_code="P-001", location="A 棟屋頂", description="施工照片封面測試")
        section = self.store.sections()[0]
        slot = self.store.slots(section["id"])[0]
        self.store.relocate_photos(imported, None, slot["id"])
        output = export_pdf(self.store, self.root / "exports" / "cover-layout.pdf")
        from pypdf import PdfReader
        reader = PdfReader(str(output))
        self.assertGreaterEqual(len(reader.pages), 2)
        cover_text = reader.pages[0].extract_text() or ""
        photo_text = reader.pages[1].extract_text() or ""
        self.assertIn("工程基本資料", cover_text)
        self.assertIn("工程編號", cover_text)
        self.assertIn("第 1 章", photo_text)

    def test_pdf_respects_selected_photos_per_page(self):
        imported, _ = self.store.import_files(
            [self.source_dir / f"photo_{index}.jpg" for index in range(3)]
        )
        extra_dir = Path(self.temp.name) / "extra"
        extra_dir.mkdir()
        for index in (3, 4):
            Image.new("RGB", (640, 480), (index * 30, 120, 80)).save(
                extra_dir / f"photo_{index}.jpg"
            )
        extra, _ = self.store.import_files(sorted(extra_dir.glob("*.jpg")))
        imported.extend(extra)
        section = self.store.sections()[0]
        slot = self.store.slots(section["id"])[0]
        self.store.relocate_photos(imported, None, slot["id"])

        from pypdf import PdfReader

        expected_pages = {1: 5, 2: 3, 4: 2, 6: 1}
        for photos_per_page, expected in expected_pages.items():
            output = self.root / "exports" / f"photos-{photos_per_page}.pdf"
            export_pdf(
                self.store,
                output,
                ExportSettings(show_cover=False, photos_per_page=photos_per_page),
            )
            self.assertEqual(expected, len(PdfReader(str(output)).pages))

    def test_pdf_starts_each_group_on_new_page(self):
        imported, _ = self.store.import_files(
            [self.source_dir / f"photo_{index}.jpg" for index in range(3)]
        )
        first_section = self.store.sections()[0]
        second_section = self.store.sections()[1]
        first_slot = self.store.slots(first_section["id"])[0]
        second_slot = self.store.slots(second_section["id"])[0]
        self.store.relocate_photos(imported[:2], None, first_slot["id"])
        self.store.relocate_photos(imported[2:], None, second_slot["id"])

        from pypdf import PdfReader

        output = self.root / "exports" / "compact-groups.pdf"
        export_pdf(
            self.store,
            output,
            ExportSettings(show_cover=False, photos_per_page=4),
        )
        self.assertEqual(2, len(PdfReader(str(output)).pages))

    def test_pdf_templates_generate(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg"])
        section = self.store.sections()[0]
        slot = self.store.slots(section["id"])[0]
        self.store.relocate_photos(imported, None, slot["id"])

        from app.exporters import PDF_TEMPLATE_OPTIONS
        from pypdf import PdfReader

        for template_key, _label, _description in PDF_TEMPLATE_OPTIONS:
            output = self.root / "exports" / f"template-{template_key}.pdf"
            export_pdf(
                self.store,
                output,
                ExportSettings(show_cover=False, pdf_template=template_key),
            )
            self.assertGreaterEqual(len(PdfReader(str(output)).pages), 1)

    def test_rotation_is_non_destructive_and_rendered(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg"])
        photo = self.store.photo(imported[0])
        original_path = self.store.root / photo["original_rel_path"]
        original_hash = self.store._sha256(original_path)
        self.store.update_photo(imported[0], rotation_degrees=90)
        rendered = self.store.rendered_photo_path(imported[0], max_side=1600)
        self.assertIsNotNone(rendered)
        self.assertEqual(90, self.store.photo(imported[0])["rotation_degrees"])
        self.assertEqual(original_hash, self.store._sha256(original_path))
        with Image.open(rendered) as image:
            self.assertEqual((480, 640), image.size)

    def test_mosaic_is_stored_in_derived_render_only(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg"])
        photo_id = imported[0]
        original = self.store.photo(photo_id)
        original_path = self.store.root / original["original_rel_path"]
        original_hash = self.store._sha256(original_path)
        self.store.update_photo(photo_id, mosaic_json="[[0.1, 0.1, 0.5, 0.5]]")
        rendered = self.store.rendered_photo_path(photo_id, max_side=640)
        self.assertIsNotNone(rendered)
        self.assertEqual(original_hash, self.store._sha256(original_path))
        self.assertEqual("[[0.1, 0.1, 0.5, 0.5]]", self.store.photo(photo_id)["mosaic_json"])

    def test_mosaic_supports_circle_and_hand_drawn_regions(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg"])
        self.store.update_photo(imported[0], mosaic_json='[{"type":"ellipse","box":[0.1,0.1,0.5,0.5]},{"type":"polygon","points":[[0.5,0.5],[0.8,0.5],[0.7,0.8]]}]')
        rendered = self.store.rendered_photo_path(imported[0], max_side=640)
        self.assertIsNotNone(rendered)

    def test_mosaic_duplicate_keeps_original_photo(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg"])
        photo_id = imported[0]
        duplicate_id = self.store.duplicate_photo_with_mosaic(photo_id, "[[0.2, 0.2, 0.4, 0.4]]", None)
        self.assertNotEqual(photo_id, duplicate_id)
        self.assertEqual("", self.store.photo(photo_id)["mosaic_json"])
        self.assertEqual("[[0.2, 0.2, 0.4, 0.4]]", self.store.photo(duplicate_id)["mosaic_json"])
        self.assertIsNotNone(self.store.rendered_photo_path(duplicate_id, max_side=320))

    def test_user_settings_are_persisted(self):
        settings_path = Path(self.temp.name) / "settings.json"
        settings = AppSettings(settings_path)
        settings.set_general(project_directory="C:/projects", export_directory="C:/reports", show_timestamp=False, show_filename=False, open_export_after=False)
        settings.save_api_profile({"name": "本機 API", "endpoint": "http://localhost:1234/v1", "api_key": "secret", "model": "demo"})
        settings.save()
        loaded = AppSettings(settings_path)
        self.assertEqual(Path("C:/reports"), loaded.export_directory)
        self.assertFalse(loaded.show_timestamp)
        self.assertFalse(loaded.show_filename)
        self.assertEqual("本機 API", loaded.active_api_profile()["name"])
        self.assertTrue(loaded.project_export_directory("屋頂/防水工程").name != "屋頂/防水工程")

    def test_five_project_data_templates_are_saved_and_switchable(self):
        settings_path = Path(self.temp.name) / "settings.json"
        settings = AppSettings(settings_path)
        self.assertEqual(5, len(settings.project_templates()))
        settings.save_project_template(2, "常用防水工程", {"company_name": "測試工程公司", "description": "固定說明"})
        settings.save()
        loaded = AppSettings(settings_path)
        self.assertEqual("常用防水工程", loaded.project_templates()[2]["name"])
        self.assertEqual("測試工程公司", loaded.load_project_template(2)["company_name"])
        self.assertEqual({}, loaded.load_project_template(4))

    def test_field_numbering_is_current_slot_only_and_rendered(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg", self.source_dir / "photo_1.jpg"])
        section = self.store.sections()[0]
        slot = self.store.slots(section["id"])[0]
        other_section = self.store.sections()[1]
        other_slot = self.store.slots(other_section["id"])[0]
        self.store.relocate_photos([imported[0], imported[1]], None, slot["id"])
        self.store.relocate_photos([imported[0]], slot["id"], other_slot["id"])
        self.store.number_slot(slot["id"], "bottom_left", "#1565C0", "96", transparent=True)
        first = self.store.placement_for_photo(imported[1], slot["id"])
        other = self.store.placement_for_photo(imported[0], other_slot["id"])
        self.assertEqual("1", first["number_text"])
        self.assertEqual(1, first["number_background_transparent"])
        self.assertEqual("", other["number_text"])
        self.assertIsNotNone(self.store.rendered_placement_path(first["id"], max_side=320))
        self.assertEqual("96", first["number_size"])
        self.store.clear_slot_numbers(slot["id"])
        self.assertEqual("", self.store.placement_for_photo(imported[1], slot["id"])["number_text"])

    def test_export_path_never_overwrites_existing_file(self):
        output_dir = self.root / "exports"
        output_dir.mkdir(exist_ok=True)
        (output_dir / "報告.pdf").touch()
        (output_dir / "報告_1.pdf").touch()
        self.assertEqual(output_dir / "報告_2.pdf", unique_output_path(output_dir, "報告", ".pdf"))

    def test_blank_project_metadata_is_omitted_from_cover_lines(self):
        lines = project_info_lines(self.store.project())
        self.assertIn("工程名稱：測試工程", lines)
        self.assertNotIn("未填寫", "\n".join(lines))
        self.assertFalse(any(line.startswith("工程編號：") for line in lines))

    def test_deleted_photos_leave_original_file_but_leave_project_lists(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_2.jpg"])
        original = self.store.root / self.store.photo(imported[0])["original_rel_path"]
        self.store.delete_photos(imported)
        self.assertTrue(original.exists())
        self.assertEqual([], self.store.photos())
        self.assertEqual([], self.store.placements_for_slot(None))

    def test_filename_can_be_hidden_from_excel_export(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg"])
        section = self.store.sections()[0]
        slot = self.store.slots(section["id"])[0]
        self.store.relocate_photos(imported, None, slot["id"])
        output = export_excel(self.store, self.root / "exports" / "no-filename.xlsx", ExportSettings(show_filename=False))
        from openpyxl import load_workbook
        workbook = load_workbook(output, read_only=True)
        self.assertNotIn("原始檔名", list(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))[0])
        workbook.close()

    def test_suggestions_can_be_accepted(self):
        section = self.store.sections()[0]
        source = self.source_dir / f"{section['name']}_照片.jpg"
        Image.new("RGB", (640, 480), (20, 20, 20)).save(source)
        imported, _ = self.store.import_files([source])
        self.assertEqual(1, self.store.create_suggestions())
        suggestions = self.store.suggestions()
        self.assertEqual(1, len(suggestions))
        self.store.accept_suggestions([suggestions[0]["id"]])
        self.assertEqual(0, len(self.store.placements_for_slot(None)))

    def test_ai_suggestions_auto_place_and_preserve_manual_override(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg"])
        target_section = self.store.sections()[0]
        target_slot = self.store.slots(target_section["id"])[0]
        manual_section = self.store.sections()[1]
        manual_slot = self.store.slots(manual_section["id"])[0]

        added, auto_placed = self.store.add_ai_suggestions([{
            "photo_id": imported[0],
            "section": target_section["name"],
            "slot": target_slot["name"],
            "confidence": 0.91,
            "reason": "測試分類",
        }])
        self.assertEqual((1, 1), (added, auto_placed))
        ai_placement = self.store.placement_for_photo(imported[0], target_slot["id"])
        self.assertIsNotNone(ai_placement)
        self.assertEqual(0, ai_placement["is_human_confirmed"])

        self.store.relocate_photos([imported[0]], target_slot["id"], manual_slot["id"])
        suggestion_id = self.store.suggestions()[0]["id"]
        self.store.accept_suggestions([suggestion_id])

        manual_placement = self.store.placement_for_photo(imported[0], manual_slot["id"])
        self.assertIsNotNone(manual_placement)
        self.assertEqual(1, manual_placement["is_human_confirmed"])
        status = self.store.db.execute("SELECT status FROM suggestions WHERE id = ?", (suggestion_id,)).fetchone()[0]
        self.assertEqual("corrected", status)

    def test_low_confidence_ai_result_stays_unclassified(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_1.jpg"])
        section = self.store.sections()[0]
        slot = self.store.slots(section["id"])[0]
        added, auto_placed = self.store.add_ai_suggestions([{
            "photo_id": imported[0],
            "section": section["name"],
            "slot": slot["name"],
            "confidence": 0.30,
        }])
        self.assertEqual((1, 0), (added, auto_placed))
        self.assertEqual(1, len(self.store.placements_for_slot(None)))
        self.assertEqual(0, len(self.store.placements_for_slot(slot["id"])))

    def test_export_check_reports_unclassified_and_passes_after_relocation(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg"])
        issues = self.store.export_check()
        self.assertEqual("error", issues[0]["severity"])
        self.assertIn("報告沒有照片", " ".join(str(issue["title"]) for issue in issues))
        section = self.store.sections()[0]
        slot = self.store.slots(section["id"])[0]
        self.store.relocate_photos(imported, None, slot["id"])
        self.assertEqual([], self.store.export_check())

    def test_export_check_warns_when_an_included_source_is_missing(self):
        imported, _ = self.store.import_files([self.source_dir / "photo_0.jpg"])
        section = self.store.sections()[0]
        slot = self.store.slots(section["id"])[0]
        self.store.relocate_photos(imported, None, slot["id"])
        photo = self.store.photo(imported[0])
        (self.store.root / photo["original_rel_path"]).unlink()
        (self.store.root / photo["preview_rel_path"]).unlink()
        issues = self.store.export_check()
        self.assertTrue(any(issue["title"] == "找不到照片檔案" for issue in issues))

    def test_exif_datetime_is_normalized_on_import(self):
        source = self.source_dir / "dated.jpg"
        image = Image.new("RGB", (320, 240), (100, 100, 100))
        exif = image.getexif()
        exif[36867] = "2026:08:20 09:08:07"
        image.save(source, exif=exif)
        imported, _ = self.store.import_files([source])
        self.assertEqual("2026-08-20 09:08:07", self.store.photo(imported[0])["captured_at"])


if __name__ == "__main__":
    unittest.main()
